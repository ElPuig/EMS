# -*- coding: utf-8 -*-

import base64
import csv
import io
import logging
from datetime import datetime

from odoo import models, fields, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class ems_applicant_import_wizard(models.TransientModel):
    _name = "ems.applicant_import_wizard"
    _description = "Applicant import wizard (GEDAC preinscription xlsx/csv)"

    file = fields.Binary(string="GEDAC file (xlsx or csv)", required=True)
    file_name = fields.Char()
    result_html = fields.Html(string="Import result", readonly=True)
    log_file = fields.Binary(string="Import log (CSV)", readonly=True)
    log_file_name = fields.Char()
    # Separate report for rows matching a still-active student (internal continuers):
    # their own data is left intact, but the destination GEDAC grants them is recorded,
    # so the enrollment proposal can reach it. Listed apart and offered as their own CSV.
    students_file = fields.Binary(string="Active students report (CSV)", readonly=True)
    students_file_name = fields.Char()

    # Columns needed to build an applicant. The many other GEDAC columns are ignored
    # (assignment metadata for other centers, "desa" duplicates, etc.).
    _REQUIRED_COLUMNS = [
        'Ident. RALC',
        'Nom',
        'Primer cognom',
        'Centre assignat',
        'Codi ensenyament assignat',
        'Nom ensenyament assignat',
        'Torn assignat',
    ]

    # GEDAC "Torn assignat" -> preinscription_shift selection. "Matí i tarda" (a
    # split-shift offer only ESO allows) maps to morning: this center always runs
    # ESO in the morning, so the afternoon half never applies.
    _SHIFT_MAP = {
        'matí': 'morning',
        'mati': 'morning',
        'tarda': 'afternoon',
        'matí i tarda': 'morning',
        'mati i tarda': 'morning',
    }

    # GEDAC "Tipus alumne" -> special_needs selection (only the NEE typologies;
    # "Ordinari" and anything else leave the field empty).
    _SEN_MAP = {
        'nee-a': 'nee_a',
        'nee-b': 'nee_b',
    }

    def action_import(self):
        center_code = self._get_center_code()
        if not center_code:
            raise UserError(_(
                "The center code is not configured. Set it in Settings > EMS Management "
                "> Center Data before importing GEDAC applicants."))

        raw = base64.b64decode(self.file)
        col_map, data_rows = self._load_table(raw)

        missing = [col for col in self._REQUIRED_COLUMNS if col not in col_map]
        if missing:
            raise UserError(
                _("The file is missing required columns:\n• %s") % "\n• ".join(missing))

        stats = {'created': 0, 'updated': 0, 'skipped': 0, 'students': 0,
                 'errors': [], 'log': [], 'student_rows': []}

        for row in data_rows:
            if not any(row):
                continue
            try:
                self._process_row(row, col_map, center_code, stats)
            except Exception as e:
                _logger.warning("Error processing GEDAC row: %s", e)
                stats['errors'].append(str(e))

        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.log_file = self._build_log_csv(stats['log'])
        self.log_file_name = 'import_gedac_%s.csv' % stamp
        if stats['student_rows']:
            self.students_file = self._build_students_csv(stats['student_rows'])
            self.students_file_name = 'gedac_alumnes_actius_%s.csv' % stamp
        self.result_html = self._build_result_html(stats)
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ems.applicant_import_wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _get_center_code(self):
        return self._norm_code(self.env.company.center_code)

    @staticmethod
    def _normalize_header(value):
        return str(value or '').strip().replace('’', "'").replace('‘', "'")

    def _load_table(self, raw):
        """Return (col_map, data_rows) from the uploaded GEDAC file, transparently
        handling both the xlsx and csv exports (same columns in both)."""
        if (self.file_name or '').lower().endswith('.csv'):
            return self._load_csv(raw)
        return self._load_xlsx(raw)

    def _load_xlsx(self, raw):
        try:
            import openpyxl
        except ImportError:
            raise UserError(_("openpyxl is required to import xlsx files."))
        wb = openpyxl.load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        for idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
            if row and any(self._normalize_header(c) == 'Ident. RALC' for c in row):
                col_map = {self._normalize_header(c): i for i, c in enumerate(row) if c}
                data_rows = list(ws.iter_rows(min_row=idx + 1, values_only=True))
                return col_map, data_rows
        raise UserError(_(
            "Could not find the header row. Make sure the file contains a column "
            "'Ident. RALC'."))

    def _load_csv(self, raw):
        # GEDAC csv exports are Latin-1 (cp1252) with a ';' delimiter; be lenient.
        text = None
        for enc in ('utf-8-sig', 'cp1252', 'latin-1'):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            text = raw.decode('latin-1', errors='replace')
        delimiter = ';'
        first_line = text.splitlines()[0] if text.splitlines() else ''
        try:
            delimiter = csv.Sniffer().sniff(first_line, delimiters=';,\t').delimiter
        except csv.Error:
            pass
        rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
        if not rows:
            raise UserError(_("The csv file is empty."))
        header = [self._normalize_header(c) for c in rows[0]]
        if 'Ident. RALC' not in header:
            raise UserError(_(
                "Could not find the header row. Make sure the file contains a column "
                "'Ident. RALC'."))
        col_map = {name: i for i, name in enumerate(header) if name}
        return col_map, rows[1:]

    def _process_row(self, row, col_map, center_code, stats):
        def get(col_name):
            idx = col_map.get(col_name)
            if idx is None:
                return None
            val = row[idx] if idx < len(row) else None
            if val is None:
                return None
            return str(val).strip() or None

        ralc = self._norm_code(get('Ident. RALC'))
        # GEDAC already splits the name: "Nom" is the (possibly multi-word) first
        # name, "Primer/Segon cognom" the surnames. Write firstname/lastname
        # directly (partner_firstname is installed) instead of a joined name, which
        # it would wrongly re-split on the first space ("Aaron Leonel Macodicomex" ->
        # firstname "Aaron"). See partner_firstname's name compute/inverse.
        firstname = get('Nom') or ''
        surname1 = get('Primer cognom') or ''
        surname2 = get('Segon cognom') or ''
        lastname = ' '.join(filter(None, [surname1, surname2]))
        name = ' '.join(filter(None, [firstname, lastname]))
        if not ralc or not name:
            return

        # Keep only applicants assigned to this center. Rows with no assignment
        # (assigned elsewhere or not granted) are skipped and logged.
        assigned_center = self._norm_code(get('Centre assignat'))
        study_code = get('Codi ensenyament assignat')
        study_name = get('Nom ensenyament assignat')
        if assigned_center != center_code or not study_code:
            stats['skipped'] += 1
            stats['log'].append({
                'accio': 'Omès', 'ralc': ralc, 'name': name, 'partner_id': False,
                'motiu': _("Not assigned to this center"), 'ts': datetime.now(),
            })
            return

        study = self._find_study(study_code, study_name)
        if not study:
            stats['skipped'] += 1
            stats['log'].append({
                'accio': 'Omès', 'ralc': ralc, 'name': name, 'partner_id': False,
                'motiu': _("Study not found: %s / %s") % (study_code, study_name),
                'ts': datetime.now(),
            })
            return

        raw_phone = get('Telèfon')
        phone, mobile = self._split_phone_mobile(self._norm_code(raw_phone))
        email = get('Correu electrònic')
        shift = self._SHIFT_MAP.get((get('Torn assignat') or '').lower())
        special_needs = self._SEN_MAP.get((get('Tipus alumne') or '').lower())
        course = self._norm_code(get('Curs'))
        course = course if course in ('1', '2', '3', '4') else False
        comment = self._build_applicant_notes(get)

        existing = self.env['res.partner'].search([('student_id', '=', ralc)], limit=1)
        if existing and existing.contact_type == 'student':
            # Internal continuer still enrolled (e.g. CFGM -> CFGS before the course
            # transition): keep the student's own data, record the destination it is
            # granted, and report it apart. Re-running the import after the transition
            # (once it is alumni) will convert it to an applicant for the new cycle.
            self._record_active_student(get, existing, study, shift, course, stats)
            return

        applicant_data = {
            'firstname': firstname or False,
            'lastname': lastname or False,
            'phone': phone,
            'mobile': mobile,
            'email': email,
            'student_id': ralc,
            'contact_type': 'applicant',
            'main_group_id': False,
            'study_id': study.id,
            'preinscription_shift': shift,
            'preinscription_course': course,
            'comment': comment or False,
        }
        # Only set when GEDAC reports a NEE typology: never clear an existing mark
        # (a re-import of an "Ordinari" row must not wipe a manually-set value).
        if special_needs:
            applicant_data['special_needs'] = special_needs

        self._get_or_create_applicant(ralc, existing, applicant_data, stats)

    def _build_applicant_notes(self, get):
        """Provenance and preinscription metadata GEDAC provides but the applicant
        model has no dedicated field for. Kept in the notes (comment) until the
        student history structure can host it (see Fase histórico)."""
        lines = []
        for label, key in [
            (_("Application number"), 'Núm. sol·licitud'),
            # NOTE: "Tipus alumne" (NEE-A/NEE-B) and "Curs" (entry course) are
            # captured in the special_needs / preinscription_course fields, not here.
            (_("Preference granted"), 'Petició atesa'),
            # Study requested as 1st choice — reveals who did not get their preferred
            # study (retention/tutoring signal) even when placed at this center.
            (_("Requested study (1st choice)"), 'Nom ensenyament'),
        ]:
            val = get(key)
            if val:
                lines.append("%s: %s" % (label, self._norm_code(val)))

        # High-level athlete flag (values come as "Sí sense valor" / "No sense valor").
        athlete = get("Esportista d'alt nivell") or ''
        if athlete.lower().startswith('s'):
            lines.append("%s: %s" % (_("High-level athlete"), _("Yes")))

        prov = [
            (_("Origin center code"), self._norm_code(get('Codi centre procedència'))),
            (_("Origin center"), get('Nom centre procedència')),
            (_("Origin study code"), self._norm_code(get('Codi ensenyament procedència'))),
            (_("Origin study"), get('Nom ensenyament procedència')),
            (_("Origin course"), self._norm_code(get('Curs procedència'))),
        ]
        prov_lines = ["%s: %s" % (label, val) for label, val in prov if val]
        if prov_lines:
            lines.append("<strong>%s</strong>" % _("Provenance"))
            lines.extend(prov_lines)

        if not lines:
            return False
        header = "[%s %s]" % (_("Import GEDAC"), datetime.now().strftime('%Y-%m-%d'))
        return '<br/>'.join([header] + lines)

    def _get_or_create_applicant(self, ralc, existing, applicant_data, stats):
        # Active students are handled apart (never reach here). Existing applicants and
        # ex-students (alumni / withdrawal) re-entering the intake become applicants again.
        if existing:
            existing.write(applicant_data)
            stats['updated'] += 1
            stats['log'].append({
                'accio': 'Actualitzat', 'ralc': ralc, 'name': existing.name,
                'partner_id': existing.id, 'motiu': '', 'ts': datetime.now(),
            })
            return existing
        applicant = self.env['res.partner'].create(applicant_data)
        stats['created'] += 1
        stats['log'].append({
            'accio': 'Creat', 'ralc': ralc, 'name': applicant.name,
            'partner_id': applicant.id, 'motiu': '', 'ts': datetime.now(),
        })
        return applicant

    def _record_active_student(self, get, student, study, shift, course, stats):
        """Record the GEDAC assignment of an already-active student found in the file,
        capturing both its current situation and the granted destination for the report.

        The student's own data (identity, group, contact details) is never overwritten:
        only the destination it is granted for next course is stored, so the enrollment
        proposal can reach it. Written as a whole, since study, shift and course are a
        single coherent assignment that each import restates.
        """
        gedac_name = ' '.join(filter(None, [
            get('Nom'), get('Primer cognom'), get('Segon cognom')]))
        shift_label = dict(self.env['res.partner']._fields['preinscription_shift'].selection).get(shift, '')
        student.write({
            'preinscription_study_id': study.id,
            'preinscription_shift': shift,
            'preinscription_course': course,
        })
        stats['students'] += 1
        stats['student_rows'].append({
            'ralc': self._norm_code(get('Ident. RALC')),
            'gedac_name': gedac_name,
            'current_name': student.name,
            'current_group': student.main_group_id.display_name or '',
            'current_study': student.study_id.display_name or '',
            'assigned_study': study.display_name,
            'shift': shift_label,
            'course': course or '',
            'email': get('Correu electrònic') or '',
            'phone': self._norm_code(get('Telèfon')) or '',
        })
        stats['log'].append({
            'accio': 'Alumne actiu', 'ralc': self._norm_code(get('Ident. RALC')),
            'name': student.name, 'partner_id': student.id,
            'motiu': _("Already an active student (internal continuer, pending course "
                       "transition): granted destination recorded"),
            'ts': datetime.now(),
        })

    def _build_students_csv(self, student_rows):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            'ralc', 'nom_gedac', 'nom_actual', 'grup_actual', 'estudi_actual',
            'estudi_assignat', 'torn_assignat', 'curs_assignat', 'correu', 'telefon',
        ])
        for r in student_rows:
            writer.writerow([
                r['ralc'], r['gedac_name'], r['current_name'], r['current_group'],
                r['current_study'], r['assigned_study'], r['shift'], r['course'],
                r['email'], r['phone'],
            ])
        return base64.b64encode(output.getvalue().encode('utf-8-sig')).decode()

    def _find_study(self, study_code, study_name):
        """Resolve the GEDAC assigned study to an ems.study.

        The GEDAC code looks like 'CFPM    IC10' (preinscription notation). Its last
        token ('IC10') matches the tail of the ems.study code ('CFGM_IC10'), so we
        match on that. Falls back to an exact name match."""
        Study = self.env['ems.study']
        token = (study_code or '').split()[-1] if study_code and study_code.split() else None
        if token:
            candidates = Study.search([('code', 'ilike', token)])
            match = candidates.filtered(
                lambda s: (s.code or '').upper().endswith(token.upper()))
            if match:
                return match[0]
        if study_name:
            byname = Study.search([('name', '=ilike', study_name)], limit=1)
            if byname:
                return byname
        return False

    def _norm_code(self, value):
        """Numeric xlsx cells arrive as floats ('8028047.0', '631078723.0'). Turn
        integer-valued floats into their plain string form for codes/ids/phones."""
        if value is None or value is False:
            return None
        text = str(value).strip()
        if not text:
            return None
        try:
            f = float(text)
            if f.is_integer():
                return str(int(f))
        except (TypeError, ValueError):
            pass
        return text

    def _split_phone_mobile(self, number, country_code='ES'):
        """Classify and canonicalize a GEDAC phone number.

        Phones arrive as plain digits, sometimes with the country code (34)
        prefixed. Valid Spanish numbers are normalized to their 9-digit national
        form (dropping the 34/+34) so every stored number looks the same, and are
        split into landline vs mobile. Numbers phonenumbers cannot validate are
        kept verbatim under phone so no data is lost (the secretary fixes them)."""
        if not number:
            return None, None
        try:
            import phonenumbers
            parsed = phonenumbers.parse(number, country_code)
            if phonenumbers.is_valid_number(parsed):
                national = phonenumbers.format_number(
                    parsed, phonenumbers.PhoneNumberFormat.NATIONAL).replace(' ', '')
                if phonenumbers.number_type(parsed) == phonenumbers.PhoneNumberType.MOBILE:
                    return None, national
                return national, None
        except Exception:
            pass
        return number, None

    def _build_log_csv(self, log_entries):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['accio', 'data_hora', 'ralc', 'id', 'nom', 'motiu'])
        for entry in log_entries:
            writer.writerow([
                entry['accio'],
                entry['ts'].strftime('%Y-%m-%d %H:%M:%S'),
                entry.get('ralc') or '',
                entry.get('partner_id') or '',
                entry.get('name') or '',
                entry.get('motiu') or '',
            ])
        return base64.b64encode(output.getvalue().encode('utf-8-sig')).decode()

    def _build_result_html(self, stats):
        errors_html = ''
        if stats['errors']:
            items = ''.join('<li>%s</li>' % e for e in stats['errors'])
            errors_html = '<p><strong>%s (%d):</strong></p><ul>%s</ul>' % (
                _("Errors"), len(stats['errors']), items)

        students_html = ''
        if stats['student_rows']:
            items = ''.join(
                '<li>%s → %s (%s)</li>' % (r['current_name'], r['assigned_study'], r['current_group'])
                for r in stats['student_rows'])
            students_html = (
                '<hr/>'
                '<p>👤 <strong>%s</strong> %d</p>'
                '<p style="color:#666;">%s</p>'
                '<ul>%s</ul>'
            ) % (
                _("Already active students (destination recorded):"), stats['students'],
                _("Internal continuers still enrolled (e.g. CFGM to CFGS before the "
                  "course transition). Their own data is kept, and the granted study, "
                  "shift and course are recorded on them: enroll them from the 'With "
                  "GEDAC assignment' filter of the enrollment proposals. Their data is "
                  "also in the 'active students' CSV below."),
                items,
            )
        return (
            '<p>✅ <strong>%s</strong> %d</p>'
            '<p>🔄 <strong>%s</strong> %d</p>'
            '<p>⏭️ <strong>%s</strong> %d</p>'
            '%s%s'
        ) % (
            _("Applicants created:"), stats['created'],
            _("Applicants updated:"), stats['updated'],
            _("Rows skipped (not assigned to this center):"), stats['skipped'],
            students_html,
            errors_html,
        )
