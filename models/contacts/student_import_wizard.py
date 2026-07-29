# -*- coding: utf-8 -*-

import base64
import csv
import io
import logging
from datetime import datetime

from markupsafe import Markup

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class EmsStudentImportWizard(models.TransientModel):
    _name = "ems.student_import_wizard"
    _description = "Student import wizard (Esfera/SAGA xlsx)"

    file = fields.Binary(string="Esfera xlsx file", required=True)
    file_name = fields.Char()
    result_html = fields.Html(string="Import result", readonly=True)
    log_file = fields.Binary(string="Import log (CSV)", readonly=True)
    log_file_name = fields.Char()

    def action_import(self):
        try:
            import openpyxl
        except ImportError:
            raise UserError(_("openpyxl is required to import xlsx files."))

        raw = base64.b64decode(self.file)
        wb = openpyxl.load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active

        header_row_idx, col_map = self._find_headers(ws)
        if header_row_idx is None:
            raise UserError(_("Could not find the header row. Make sure the xlsx file contains a column 'Grup Classe'."))

        missing = self._check_required_columns(col_map)
        if missing:
            raise UserError(_(
                "The file is missing required columns:\n• %(columns)s",
                columns="\n• ".join(missing),
            ))

        stats = {'created': 0, 'updated': 0, 'errors': [], 'log': []}

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not any(row):
                continue
            try:
                self._process_row(row, col_map, stats)
            except Exception as e:
                _logger.warning("Error processing row: %s", e)
                stats['errors'].append(str(e))

        self.log_file = self._build_log_csv(stats['log'])
        self.log_file_name = f"import_esfera_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        self.result_html = self._build_result_html(stats)
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ems.student_import_wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _find_headers(self, ws):
        def normalize(s):
            return str(s or '').strip().replace('’', "'").replace('‘', "'")
        for idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
            if row and any(normalize(c) == 'Grup Classe' for c in row):
                col_map = {normalize(c): i for i, c in enumerate(row) if c}
                return idx, col_map
        return None, {}

    _REQUIRED_COLUMNS = [
        'Grup Classe',
        'Nom',
        'Primer Cognom',
        'Segon Cognom',
        'Identificador de l\'alumne/a',
        'Número de document d\'identitat',
        'Tipus de document d\'identitat',
        'Data naixement',
        'Nacionalitat',
        'País naixement',
        'Telèfon',
        'Correu electrònic',
        'Tipus de via',
        'Nom via',
        'Número',
        'Codi postal',
        'Municipi de residència',
        'Província de residència',
        'País de residència',
        'Tutor 1 - nom',
        'Tutor 1 - doc. identitat',
        'Tutor 1 - municipi',
        'Tutor 1 - provincia',
        'Tutor 1 - país',
        'Tutor 1 - CP',
        'Tutor 2 - nom',
        'Tutor 2 - doc. identitat',
        'Tutor 2 - municipi',
        'Tutor 2 - provincia',
        'Tutor 2 - país',
        'Tutor 2 - CP',
        'Contacte 1er tutor alumne - Valor',
        'Contacte 1er tutor alumne - Observacions',
        'Contacte 2on tutor alumne - Valor',
        'Contacte 2on tutor alumne - Observacions',
    ]

    # Columns where Esfera may export with or without trailing space — check at least one variant
    _REQUIRED_COLUMNS_VARIANTS = [
        ('Tutor 1 - 1r cognom', 'Tutor 1 - 1r cognom '),
        ('Tutor 2 - 1r cognom', 'Tutor 2 - 1r cognom '),
    ]

    def _check_required_columns(self, col_map):
        missing = [col for col in self._REQUIRED_COLUMNS if col not in col_map]
        for variants in self._REQUIRED_COLUMNS_VARIANTS:
            if not any(v in col_map for v in variants):
                missing.append(variants[0])
        return missing

    def _col_get(self, row, col_map, col_name):
        """Read one Esfera column by its exact header name, or None if the
        column is absent from this file or the cell is empty for this row.
        Shared by _process_row (student columns) and _process_tutor (Tutor
        N - prefixed columns)."""
        idx = col_map.get(col_name)
        if idx is None:
            return None
        val = row[idx] if idx < len(row) else None
        return str(val).strip() if val is not None else None

    def _process_row(self, row, col_map, stats):
        def get(col_name):
            return self._col_get(row, col_map, col_name)

        # Group — search by external_id (Esfera code)
        # Normalize whitespace: Esfera sometimes uses multiple spaces (e.g. "CFPM    IC10201")
        esfera_code = ' '.join((get('Grup Classe') or '').split()) or None
        if not esfera_code:
            return
        group = self.env['ems.group'].search([('external_id', '=', esfera_code)], limit=1)

        # Student name
        firstname = get('Nom') or ''
        surname1 = get('Primer Cognom') or ''
        surname2 = get('Segon Cognom') or ''
        name = ' '.join(filter(None, [firstname, surname1, surname2]))
        if not name:
            return

        # Documents
        doc_nums = get('Número de document d\'identitat')
        doc_types = get('Tipus de document d\'identitat') or ''
        docs = self._parse_documents(doc_nums, doc_types)

        # Personal data
        ralc = get('Identificador de l\'alumne/a')
        birth_str = get('Data naixement')
        birth_date = self._parse_date(birth_str)
        citizenship = self._find_country(get('Nacionalitat'))
        birth_country = self._find_country(get('País naixement'))
        raw_phone = get('Telèfon') or get('Contacte alumne - Telèfon')
        student_phone, student_mobile = self._split_phone_mobile(raw_phone)
        email = get('Correu electrònic') or get('Contacte alumne - Correu electrònic')

        # Address
        street = self._build_street(
            get('Tipus de via'), get('Nom via'), get('Número'),
            get('Bloc'), get('Escala'), get('Planta'), get('Porta'),
            get('Resta de dades de l\'adreça')
        )
        city = get('Municipi de residència')
        zip_code = get('Codi postal')
        country = self._find_country(get('País de residència'))
        state = self._find_state(get('Província de residència'), country.id if country else False)

        # Extra notes
        comment = self._build_student_notes(get, esfera_code, group)

        student_data = {
            'name': name,
            'contact_type': 'student',
            'birth_date': birth_date,
            'document_id': docs.get('DNI') or docs.get('NIE'),
            'passport_id': docs.get('PASS') or docs.get('Passaport'),
            'medical_id': docs.get('TIS'),
            'phone': student_phone,
            'mobile': student_mobile,
            'email': email,
            'street': street,
            'city': city,
            'zip': zip_code,
            'country_id': country.id if country else False,
            'state_id': state.id if state else False,
            'citizenship_id': citizenship.id if citizenship else False,
            'birth_country_id': birth_country.id if birth_country else False,
            'comment': comment or False,
            # Re-admits an ex-student (alumni/withdrawal) archived on exit: without
            # this, an existing-but-inactive match is written but stays archived.
            'active': True,
        }
        if group:
            student_data['main_group_id'] = group.id
        if ralc:
            student_data['student_id'] = ralc

        student = self._get_or_create_student(ralc, student_data, stats)

        # Tutors
        for prefix in ['Tutor 1', 'Tutor 2']:
            self._process_tutor(row, col_map, prefix, student, stats)

    # NOTE: the note labels below are deliberately kept in Catalan, untranslated,
    # even though they end up in a user-visible Notes field. They are a verbatim
    # echo of Esfera/SAGA's own official Catalan field names (the Catalan
    # education administration's system of record) — translating them would
    # weaken traceability back to "this is exactly what Esfera exported for this
    # student", which is the whole point of keeping them. See student_import_wizard.md.
    def _build_student_notes(self, get, esfera_code, group):
        lines = []
        if not group:
            lines.append(f"Grup Classe (SAGA): {esfera_code}")
        for label, key in [
            ('Província de naixement', 'Província naixement'),
            ('Municipi de naixement', 'Municipi naixement'),
            ('Localitat de residència', 'Localitat de residència'),
            ('Alumne tutelat legalment', 'Alumne tutelat legalment'),
            ('Alumne emancipat legalment', 'Alumne emancipat legalment'),
            ('Alumne amb custòdia compartida en dos domicilis', 'Alumne amb custòdia compartida en dos domicilis'),
            ('Contacte altres - Tipus', 'Contacte altres alumne - Tipus'),
            ('Contacte altres - Valor', 'Contacte altres alumne - Valor'),
            ('Contacte altres - Observacions', 'Contacte altres alumne - Observacions'),
            ('Contacte propis - Observacions', 'Contacte propis alumne - Observacions'),
            ('Observacions', 'Observacions'),
        ]:
            val = get(key)
            if val and val.lower() not in ('no', 'false', ''):
                lines.append(f"{label}: {val}")
        return '<br/>'.join(lines) if lines else False

    def _get_or_create_student(self, ralc, data, stats):
        existing = False
        if ralc:
            existing = self.env['res.partner'].with_context(active_test=False).search(
                [('student_id', '=', ralc)], limit=1)
        if existing:
            existing.write(data)
            stats['updated'] += 1
            stats['log'].append({'tipus': 'Alumne', 'accio': 'Actualitzat', 'partner_id': existing.id, 'ts': datetime.now()})
            return existing
        student = self.env['res.partner'].create(data)
        stats['created'] += 1
        stats['log'].append({'tipus': 'Alumne', 'accio': 'Creat', 'partner_id': student.id, 'ts': datetime.now()})
        return student

    def _process_tutor(self, row, col_map, prefix, student, stats):
        def get(col_name):
            return self._col_get(row, col_map, col_name)

        nom = get(f'{prefix} - nom')
        if not nom:
            return

        # Note: Esfera exports '1r cognom ' with trailing space
        surname1 = get(f'{prefix} - 1r cognom ') or get(f'{prefix} - 1r cognom') or ''
        surname2 = get(f'{prefix} - 2n cognom') or ''
        full_name = ' '.join(filter(None, [nom, surname1, surname2]))

        doc_num = get(f'{prefix} - doc. identitat')
        docs = self._parse_documents(doc_num, '')

        tutor_num = '1er' if '1' in prefix else '2on'
        contact_raw = get(f'Contacte {tutor_num} tutor alumne - Valor')
        raw_phone, email = self._parse_contact_value(contact_raw)
        phone, mobile = self._split_phone_mobile(raw_phone)
        observacio = get(f'Contacte {tutor_num} tutor alumne - Observacions')

        street = self._build_street(
            get(f'{prefix} - tipus via'), get(f'{prefix} - nom via'),
            get(f'{prefix} - número'), get(f'{prefix} - bloc'),
            get(f'{prefix} - escala'), get(f'{prefix} - planta'),
            get(f'{prefix} - porta'), get(f'{prefix} - resta de dades de l\'adreça')
        )
        city = get(f'{prefix} - municipi')
        zip_code = get(f'{prefix} - CP')
        country = self._find_country(get(f'{prefix} - país'))
        state = self._find_state(get(f'{prefix} - provincia'), country.id if country else False)

        # Extra notes for tutor — same Catalan-verbatim rationale as _build_student_notes.
        tutor_notes = []
        for label, key in [
            ('Localitat', f'{prefix} - localitat'),
            ('Destinatari correspondència', f'{prefix} - destinatari correspondència'),
            ('Persona jurídica', f'{prefix} - persona jurídica'),
            ('Rep notificacions', f'{prefix} - contactes: rebre notificacions'),
        ]:
            val = get(key)
            if val and val.lower() not in ('no', 'false', ''):
                tutor_notes.append(f"{label}: {val}")
        if prefix == 'Tutor 2':
            shared = get('Tutors comparteixen domicili')
            if shared and shared.lower() not in ('no', 'false', ''):
                tutor_notes.append(f"Comparteix domicili amb Tutor 1: {shared}")

        family, accio = self._get_or_create_family(
            full_name, doc_num, phone, mobile, email,
            {'street': street, 'city': city, 'zip': zip_code,
             'country_id': country.id if country else False,
             'state_id': state.id if state else False,
             'comment': '<br/>'.join(tutor_notes) if tutor_notes else False}
        )
        if not family:
            return
        stats['log'].append({'tipus': 'Familiar', 'accio': accio, 'partner_id': family.id, 'ts': datetime.now()})

        relation_type, is_fallback = self._deduce_relation_type(observacio)
        if is_fallback and observacio:
            note = f"[Import Esfera] Relació {prefix}: '{observacio}' (assignada com a Tutor per defecte)"
            student.comment = f"{student.comment}<br/>{note}".strip() if student.comment else note

        self._link_family_to_student(family, student, relation_type)

    def _get_or_create_family(self, name, doc_num, phone, mobile, email, address_data):
        """Find or create the family contact for a tutor row.

        KNOWN LIMITATION: dedup only matches on doc_num (document_id/passport_id).
        A tutor row with no document number always creates a new family partner —
        there is no name/phone/email fallback match, so re-importing the same file
        for a documentless tutor creates one duplicate family contact per import
        run. Flagged, not fixed, in this DTON pass — see student_import_wizard.md.
        """
        if not name:
            return False, None

        domain = [('contact_type', '=', 'family')]
        existing = False
        if doc_num:
            existing = (
                self.env['res.partner'].search(domain + [('document_id', '=', doc_num)], limit=1)
                or self.env['res.partner'].search(domain + [('passport_id', '=', doc_num)], limit=1)
            )
        if existing:
            update_vals = dict(address_data)
            if phone:
                update_vals['phone'] = phone
            if mobile:
                update_vals['mobile'] = mobile
            if email:
                update_vals['email'] = email
            existing.write(update_vals)
            return existing, 'Actualitzat'

        vals = dict(address_data, **{
            'name': name,
            'contact_type': 'family',
            'document_id': doc_num,
            'phone': phone,
            'mobile': mobile,
            'email': email,
        })
        return self.env['res.partner'].create(vals), 'Creat'

    def _link_family_to_student(self, family, student, relation_type):
        existing = self.env['res.partner.relation'].search([
            ('left_partner_id', '=', family.id),
            ('right_partner_id', '=', student.id),
        ], limit=1)
        if not existing:
            self.env['res.partner.relation'].create({
                'left_partner_id': family.id,
                'type_id': relation_type.id,
                'right_partner_id': student.id,
            })

    def _deduce_relation_type(self, text):
        t = (text or '').lower()
        is_fallback = False
        if 'mare' in t or 'madre' in t:
            rel = self.env.ref('ems.relation_type_mother')
        elif 'pare' in t or 'padre' in t:
            rel = self.env.ref('ems.relation_type_father')
        elif 'àvia' in t or 'avia' in t or 'abuela' in t:
            rel = self.env.ref('ems.relation_type_grandmother')
        elif 'avi' in t or 'abuelo' in t:
            rel = self.env.ref('ems.relation_type_grandfather')
        elif 'tieta' in t or 'tia' in t or 'oncle' in t or 'tio' in t:
            rel = self.env.ref('ems.relation_type_uncle_aunt')
        elif 'germana' in t or 'hermana' in t or 'germà' in t or 'hermano' in t:
            rel = self.env.ref('ems.relation_type_sibling')
        else:
            rel = self.env.ref('ems.relation_type_tutor')
            is_fallback = True
        return rel, is_fallback

    def _parse_documents(self, numbers_str, types_str):
        if not numbers_str:
            return {}
        numbers = [n.strip() for n in numbers_str.split(' - ')]
        types = [t.strip() for t in (types_str or '').split(' - ')]
        return dict(zip(types, numbers))

    def _parse_date(self, value):
        if not value:
            return False
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
            try:
                return datetime.strptime(str(value), fmt).date()
            except ValueError:
                continue
        return False

    def _build_street(self, tipus_via, nom_via, numero, bloc, escala, planta, porta, resta):
        parts = [tipus_via, nom_via, numero, bloc, escala, planta, porta, resta]
        return ' '.join(p for p in parts if p and str(p).strip()) or False

    def _parse_contact_value(self, raw):
        if not raw:
            return None, None
        parts = [p.strip() for p in raw.split(' - ')]
        phone = parts[0] if parts else None
        email = parts[1] if len(parts) > 1 else None
        return phone, email

    def _split_phone_mobile(self, number, country_code='ES'):
        if not number:
            return None, None
        try:
            import phonenumbers
            parsed = phonenumbers.parse(number, country_code)
            if phonenumbers.number_type(parsed) == phonenumbers.PhoneNumberType.MOBILE:
                return None, number
            return number, None
        except Exception:
            return number, None

    def _find_country(self, name):
        if not name:
            return False
        return self.env['res.country'].with_context(lang='ca_ES').search(
            [('name', 'ilike', name)], limit=1
        )

    def _find_state(self, name, country_id):
        if not name or not country_id:
            return False
        return self.env['res.country.state'].with_context(lang='ca_ES').search(
            [('name', 'ilike', name), ('country_id', '=', country_id)], limit=1
        )

    def _build_log_csv(self, log_entries):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            'tipus', 'accio', 'data_hora', 'id', 'nom', 'document_id',
            'email', 'telèfon', 'mòbil', 'grup', 'vinculat_a', 'vinculat_als_ids',
        ])
        for entry in log_entries:
            partner = self.env['res.partner'].browse(entry['partner_id'])
            if entry['tipus'] == 'Alumne':
                relations = self.env['res.partner.relation'].search([('right_partner_id', '=', partner.id)])
                linked = relations.mapped('left_partner_id')
                grup = partner.main_group_id.display_name if partner.main_group_id else ''
            else:
                relations = self.env['res.partner.relation'].search([('left_partner_id', '=', partner.id)])
                linked = relations.mapped('right_partner_id')
                grup = ''
            writer.writerow([
                entry['tipus'],
                entry['accio'],
                entry['ts'].strftime('%Y-%m-%d %H:%M:%S'),
                partner.id,
                partner.name,
                partner.document_id or '',
                partner.email or '',
                partner.phone or '',
                partner.mobile or '',
                grup,
                ', '.join(linked.mapped('name')),
                ', '.join(str(p.id) for p in linked),
            ])
        return base64.b64encode(output.getvalue().encode('utf-8-sig')).decode()

    def _build_result_html(self, stats):
        errors_html = ''
        if stats['errors']:
            errors_html = Markup('<p><strong>{}</strong></p>{}').format(
                _("Errors (%(count)s):", count=len(stats['errors'])),
                self.env['ems.base'].build_html_list(stats['errors']),
            )
        return Markup(
            '<p>✅ <strong>{created_label}</strong> {created}</p>'
            '<p>🔄 <strong>{updated_label}</strong> {updated}</p>'
            '{errors_html}'
        ).format(
            created_label=_("Students created:"), created=stats['created'],
            updated_label=_("Students updated:"), updated=stats['updated'],
            errors_html=errors_html,
        )
