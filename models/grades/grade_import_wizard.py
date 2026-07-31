# -*- coding: utf-8 -*-

import base64
import csv
import io
import logging
import re
from datetime import datetime

from odoo import models, fields, api, _
from odoo.exceptions import UserError

from .grade_session import grade_round_selection

_logger = logging.getLogger(__name__)

# Columns of the "Notes Flat" sheet, in order.
_FLAT_HEADERS = ["idAlumne", "nom Alumne", "Codi Mòdul", "Nom Mòdul", "Codi", "Nom", "Tipus", "Subtipus", "Nota"]

# Header cells of the pivoted "Notes" sheet that are not grade columns.
_PIVOT_SKIP_HEADERS = {"idalumne", "nom", "n. convocatoria", "n. convocatòria", "provisional"}

# A grade column header in the pivoted "Notes" sheet: "<module>_<NN><RA|EM>".
_PIVOT_OUTCOME_RE = re.compile(r"^(?P<mod>.+)_(?P<sub>\d{2})(?P<kind>RA|EM)$")

# Aggregate "module" codes that Esfera exports for higher cycles (overall final grade and
# university-access grade). They are not real subjects, so they are skipped instead of reported
# as errors. Add any other export-only aggregate codes here.
_SKIP_MODULE_CODES = {"QFINAL", "QUNIVERSITAT"}


class ems_grade_import_wizard(models.TransientModel):
	_name = "ems.grade_import_wizard"
	_description = "Grade import wizard (Esfera xlsx)"

	round = fields.Selection(string="Evaluation", selection=grade_round_selection, required=True)
	file = fields.Binary(string="Esfera xlsx file", required=True)
	file_name = fields.Char()
	create_missing_enrollments = fields.Boolean(
		string="Create missing enrollments", default=False,
		help="Enroll students that are graded in a module they are not enrolled in, so the grade can "
			 "be imported instead of being discarded. Any informed grade counts, numeric or textual "
			 "(PDT, NP, CV...); a module left entirely blank and modules without an evaluation "
			 "session are never enrolled, and an optional module only when the group has a single "
			 "optional subject, since otherwise there is no way to tell which.")
	result_html = fields.Html(string="Import result", readonly=True)
	log_file = fields.Binary(string="Import log (CSV)", readonly=True)
	log_file_name = fields.Char()

	def action_import(self):
		self.ensure_one()
		try:
			import openpyxl
		except ImportError:
			raise UserError(_("openpyxl is required to import xlsx files."))

		raw = base64.b64decode(self.file)
		wb = openpyxl.load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)

		rows = self._read_rows(wb)
		if not rows:
			raise UserError(_("The file has no gradeable rows. Expected a 'Notes Flat' or 'Notes' sheet."))

		stats = {
			"ra": 0, "em": 0, "mp": 0, "locked": 0, "enrollments": 0,
			"warnings": [], "errors": [], "log": [],
		}

		context = self._build_context(rows, stats)
		self._apply_rows(rows, context, stats)

		self.log_file = self._build_log_csv(stats["log"])
		self.log_file_name = "import_grades_%s.csv" % datetime.now().strftime("%Y%m%d_%H%M%S")
		self.result_html = self._build_result_html(stats)
		return {
			"type": "ir.actions.act_window",
			"res_model": self._name,
			"res_id": self.id,
			"view_mode": "form",
			"target": "new",
		}

	# --- Reading: both sheets are normalised to (idalu, codi_modul, tipus, subtipus, nota) tuples ---

	def _read_rows(self, wb):
		if "Notes Flat" in wb.sheetnames:
			return self._read_flat(wb["Notes Flat"])
		if "Notes" in wb.sheetnames:
			return self._read_pivot(wb["Notes"])
		return []

	def _read_flat(self, ws):
		rows = []
		it = ws.iter_rows(values_only=True)
		header = next(it, None)  # first row is the header
		if not header:
			return rows
		# Map the expected flat headers to their column index (tolerant to ordering).
		idx = {str(cell or "").strip(): i for i, cell in enumerate(header)}
		c_idalu = idx.get("idAlumne")
		c_mod = idx.get("Codi Mòdul")
		c_tipus = idx.get("Tipus")
		c_sub = idx.get("Subtipus")
		c_nota = idx.get("Nota")
		if None in (c_idalu, c_mod, c_tipus, c_sub, c_nota):
			raise UserError(_("The 'Notes Flat' sheet is missing required columns (idAlumne, Codi Mòdul, Tipus, Subtipus, Nota)."))
		for row in it:
			if not row or row[c_idalu] in (None, ""):
				continue
			rows.append((
				self._clean_idalu(row[c_idalu]),
				str(row[c_mod] or "").strip(),
				str(row[c_tipus] or "").strip(),
				str(row[c_sub] or "").strip(),
				row[c_nota],
			))
		return rows

	def _read_pivot(self, ws):
		# The header row is the one containing "idAlumne"; columns before it may be a merged title row.
		grid = list(ws.iter_rows(values_only=True))
		header_idx = None
		for i, row in enumerate(grid[:20]):
			if row and any(str(c or "").strip().lower() == "idalumne" for c in row):
				header_idx = i
				break
		if header_idx is None:
			return []
		header = grid[header_idx]
		# Classify each column: (kind, codi_modul, subtipus) or None to skip.
		col_spec = {}
		id_col = None
		for i, cell in enumerate(header):
			label = str(cell or "").strip()
			low = label.lower()
			if low == "idalumne":
				id_col = i
				continue
			if not label or low in _PIVOT_SKIP_HEADERS:
				continue
			match = _PIVOT_OUTCOME_RE.match(label)
			if match:
				col_spec[i] = (match.group("kind"), match.group("mod"), match.group("sub"))
			else:
				# A bare module code column is the module's final grade (MP).
				col_spec[i] = ("MP", label, "MP")
		if id_col is None:
			return []
		rows = []
		for row in grid[header_idx + 1:]:
			if not row or id_col >= len(row) or row[id_col] in (None, ""):
				continue
			idalu = self._clean_idalu(row[id_col])
			for i, (kind, mod, sub) in col_spec.items():
				if i >= len(row):
					continue
				rows.append((idalu, mod, kind, sub, row[i]))
		return rows

	@api.model
	def _clean_idalu(self, value):
		# idAlumne comes as an integer in xlsx; res.partner.student_id is a Char.
		if isinstance(value, float) and value.is_integer():
			value = int(value)
		return str(value).strip()

	# --- Preprocessing: resolve students, derive group(s), load sessions and build indexes ---

	def _build_context(self, rows, stats):
		idalus = {row[0] for row in rows if row[0]}
		partners = self.env["res.partner"].search([
			("contact_type", "=", "student"),
			("student_id", "in", list(idalus)),
		])
		student_by_idalu = {p.student_id: p for p in partners}

		main_groups = partners.mapped("main_group_id")
		if len(main_groups) > 1:
			stats["warnings"].append(_("The file spans several groups: %s.") % ", ".join(main_groups.mapped("name")))

		# A module is not always taught in the student's own group: split groups have their own
		# ems.group, and a repeater takes a lower-year module with that year's group. The enrollment
		# carries the group where the student actually attends, and that is where their grade line
		# lives, so the sessions are looked up through the enrollments and not only through the main
		# group. Relying on the main group alone made the result depend on who else was in the file:
		# the session showed up only if some other student happened to have it as their main group.
		enrollments = self.env["ems.enrollment"].search([("student_id", "in", partners.ids)])
		groups = main_groups | enrollments.mapped("group_id")

		sessions = self.env["ems.grade_session"].search([
			("group_id", "in", groups.ids),
			("round", "=", self.round),
		])
		self._warn_on_split_enrollments(enrollments, stats)
		# Candidate subjects for module matching, and the session each (group, subject) is graded in.
		subject_by_code = {}
		outcome_by_code = {}
		session_by_subject = {}
		for session in sessions:
			subject = session.subject_id
			subject_by_code.setdefault(subject.code, subject)
			session_by_subject.setdefault((session.group_id.id, subject.id), session)
			for outcome in subject.outcome_ids:
				outcome_by_code[outcome.code] = outcome

		# Enrolling adds the student's grade lines, so it must happen before the line indexes below
		# are built for those lines to be picked up by this same import.
		if self.create_missing_enrollments:
			self._create_missing_enrollments(rows, student_by_idalu, subject_by_code, session_by_subject, stats)
			sessions.invalidate_recordset(["grade_outcome_line_ids", "grade_subject_line_ids"])

		# O(1) line indexes.
		outcome_line = {}
		subject_line = {}
		# Optional modules ("MP OPTx") are named/coded differently in each centre, so their Esfera code
		# never matches the EMS code (e.g. Esfera "OPT2" vs EMS "OPT1" for the same cycle). They are
		# resolved by the optional subject the student is actually enrolled in (one per study).
		optional_by_student = {}
		for session in sessions:
			subject = session.subject_id
			is_optional = subject.code.upper().startswith("OPT")
			for line in session.grade_outcome_line_ids:
				outcome_line[(line.student_id.id, line.outcome_id.id)] = line
			for line in session.grade_subject_line_ids:
				subject_line[(line.student_id.id, subject.id)] = line
				if is_optional:
					optional_by_student[line.student_id.id] = subject
		return {
			"student_by_idalu": student_by_idalu,
			"subject_by_code": subject_by_code,
			"outcome_by_code": outcome_by_code,
			"outcome_line": outcome_line,
			"subject_line": subject_line,
			"optional_by_student": optional_by_student,
		}

	def _warn_on_split_enrollments(self, enrollments, stats):
		"""Report students enrolled in the same module in more than one group.

		Widening the session lookup to every group the student is enrolled in makes such a student
		have a grade line in two sessions at once, and the line indexes below are keyed by student
		and outcome, with no group: one of the two lines would silently win. It is a data problem
		(a module is attended in one group), so it is reported instead of guessed.
		"""
		groups_by_pair = {}
		for enrollment in enrollments:
			pair = (enrollment.student_id, enrollment.subject_id)
			groups_by_pair.setdefault(pair, self.env["ems.group"])
			groups_by_pair[pair] |= enrollment.group_id
		for (student, subject), groups in groups_by_pair.items():
			if len(groups) > 1:
				stats["warnings"].append(
					_("%s is enrolled in module '%s' in several groups (%s): the grade may be written "
					  "to any of their sessions. Leave a single enrollment per module.")
					% (student.display_name, subject.code, ", ".join(groups.mapped("name"))))

	def _create_missing_enrollments(self, rows, student_by_idalu, subject_by_code, session_by_subject, stats):
		"""Enroll students that are graded in a module they are not enrolled in.

		Any informed grade counts, numeric or textual: a textual one is a grade too, not the absence
		of one. "PDT"/"NP" say the module is not passed and "CV" (convalidated) says it is, but all
		of them state that the module is part of the student's record. What does not count is a
		module left entirely blank, which is how Esfera lists the cycle's modules a student does not
		take. Modules with no session for the student's group are skipped too (nothing to grade into).

		Optional modules cannot be matched by code (Esfera's "OPT2" against this centre's own code),
		and what normally resolves them - the student's enrollment - is precisely what is missing
		here. They are therefore resolved by elimination: if the group has exactly one optional
		subject being graded this round, that is unambiguously the one; with two or more there is no
		way to tell which, so nothing is created and the case is reported.
		"""
		# Which (student, module) pairs of the file carry at least one informed grade, in file order.
		pairs = []
		seen = set()
		graded_pairs = set()
		for idalu, codi_modul, _tipus, _subtipus, nota in rows:
			pair = (idalu, codi_modul)
			if pair not in seen:
				seen.add(pair)
				pairs.append(pair)
			if nota is not None and str(nota).strip():
				graded_pairs.add(pair)

		# The optional subjects being graded this round, per group, for the resolution by elimination.
		optional_sessions = {}
		for (group_id, _subject_id), session in session_by_subject.items():
			if session.subject_id.code.upper().startswith("OPT"):
				optional_sessions.setdefault(group_id, []).append(session)

		candidates = []
		for idalu, codi_modul in pairs:
			if (idalu, codi_modul) not in graded_pairs:
				continue
			if codi_modul.upper() in _SKIP_MODULE_CODES:
				continue
			student = student_by_idalu.get(idalu)
			if not student:
				continue  # already reported as an unknown student by _apply_rows
			if codi_modul.upper().startswith("OPT"):
				group_optionals = optional_sessions.get(student.main_group_id.id, [])
				if len(group_optionals) != 1:
					if len(group_optionals) > 1:
						stats["warnings"].append(
							_("%s is graded in optional module '%s', but group %s has %d optional subjects: "
							  "cannot tell which one to enroll them in.")
							% (student.display_name, codi_modul, student.main_group_id.display_name,
							   len(group_optionals)))
					continue
				session = group_optionals[0]
				subject = session.subject_id
			else:
				# The optional mapping is irrelevant here: optional modules are handled above.
				subject = self._resolve_subject(codi_modul, student, {
					"subject_by_code": subject_by_code, "optional_by_student": {},
				})
				session = session_by_subject.get((student.main_group_id.id, subject.id)) if subject else None
			if not session:
				continue  # reported as a missing session by _apply_rows
			candidates.append((student, subject, session))
		if not candidates:
			return

		# One search for every candidate: an enrollment in another group means the student is placed
		# elsewhere for this module, which is an anomaly to review rather than one to fix by adding
		# a second enrollment.
		enrollment_model = self.env["ems.enrollment"]
		existing = enrollment_model.search([
			("student_id", "in", [student.id for student, _subject, _session in candidates]),
			("subject_id", "in", [subject.id for _student, subject, _session in candidates]),
		])
		enrolled = {(enrollment.student_id.id, enrollment.subject_id.id): enrollment for enrollment in existing}

		to_create = []
		for student, subject, session in candidates:
			enrollment = enrolled.get((student.id, subject.id))
			if enrollment:
				if enrollment.group_id != session.group_id:
					stats["warnings"].append(
						_("%s is graded in '%s' with group %s but is enrolled in group %s: enrollment left untouched.")
						% (student.display_name, subject.code, session.group_id.display_name,
						   enrollment.group_id.display_name))
				continue
			enrolled[(student.id, subject.id)] = True  # guard against duplicates within this run
			to_create.append((student, subject, session))

		if not to_create:
			return
		enrollment_model.create([{
			"student_id": student.id, "group_id": session.group_id.id, "subject_id": subject.id,
		} for student, subject, session in to_create])
		for student, subject, session in to_create:
			# ems.enrollment.create() already fills the open sessions, but this import may target a
			# session in the board/final state (an administrator can still write to it), so the lines
			# are added explicitly. _ems_add_student_lines is idempotent.
			session._ems_add_student_lines(student)
			stats["enrollments"] += 1
			stats["log"].append({
				"tipus": "ENROLLMENT", "accio": "CREATED", "idalu": student.student_id,
				"alumne": student.display_name, "modul": subject.code, "codi": session.group_id.display_name,
				"nota": "",
			})

	def _resolve_subject(self, codi_modul, student, ctx):
		# Optional modules: the Esfera and EMS codes differ by design, so map straight to the optional
		# subject the student is enrolled in, ignoring the code.
		if codi_modul.upper().startswith("OPT"):
			return ctx["optional_by_student"].get(student.id)
		by_code = ctx["subject_by_code"]
		if codi_modul in by_code:
			return by_code[codi_modul]
		# Strip the trailing cycle token (e.g. "0156_IC10" -> "0156").
		stripped = codi_modul.rsplit("_", 1)[0]
		return by_code.get(stripped)

	# --- Applying: two passes (RA/EM first so MP can read the recomputed final) ---

	def _apply_rows(self, rows, ctx, stats):
		mp_rows = []
		for idalu, codi_modul, tipus, subtipus, nota in rows:
			# Skip export-only aggregate columns (overall / university-access final grade): they are not
			# subjects, so they must not be reported as missing sessions.
			if codi_modul.upper() in _SKIP_MODULE_CODES:
				continue
			student = ctx["student_by_idalu"].get(idalu)
			if not student:
				self._log_error(stats, idalu, codi_modul, tipus, _("Student not found (idAlumne %s).") % idalu)
				continue
			subject = self._resolve_subject(codi_modul, student, ctx)
			if not subject:
				self._log_error(stats, idalu, codi_modul, tipus,
					_("No evaluation session for module '%s' in this evaluation.") % codi_modul)
				continue

			if tipus == "RA":
				self._apply_ra(student, subject, subtipus, nota, ctx, stats)
			elif tipus == "EM":
				self._apply_em(student, subject, nota, ctx, stats)
			elif tipus == "MP":
				mp_rows.append((student, subject, nota))
			# Unknown Tipus values are ignored silently.

		for student, subject, nota in mp_rows:
			self._apply_mp(student, subject, nota, ctx, stats)

	def _apply_ra(self, student, subject, subtipus, nota, ctx, stats):
		code = "%s_%s%s" % (subject.code, subtipus, "RA")
		outcome = ctx["outcome_by_code"].get(code)
		if not outcome:
			self._log_error(stats, student.student_id, subject.code, "RA",
				_("Outcome '%s' not found for subject '%s'.") % (code, subject.code))
			return
		line = ctx["outcome_line"].get((student.id, outcome.id))
		if not line:
			self._log_error(stats, student.student_id, subject.code, "RA",
				_("No grade line for student in outcome '%s' (not enrolled or session not filled).") % code)
			return
		score, scored = self._coerce_score(nota)
		vals = {"is_scored": scored}
		if scored:
			vals["score"] = score
		if line.is_locked:
			# The official grade overwrites a locked outcome: release the lock on this line only, so the
			# imported grade is shown without the padlock. Earlier rounds are left intact (their history
			# is preserved); the lock recomputes from them again for any future round.
			stats["locked"] += 1
			vals["is_lock_released"] = True
		line.with_context(ems_grade_import_bypass_lock=True).write(vals)
		stats["ra"] += 1
		self._log(stats, "RA", student, subject, code, nota)

	def _apply_em(self, student, subject, nota, ctx, stats):
		line = ctx["subject_line"].get((student.id, subject.id))
		if not line:
			self._log_error(stats, student.student_id, subject.code, "EM",
				_("No subject grade line for student (not enrolled or session not filled)."))
			return
		score, scored = self._coerce_score(nota)
		vals = {"external_is_scored": scored}
		if scored:
			vals["external_score"] = score
		line.write(vals)
		stats["em"] += 1
		self._log(stats, "EM", student, subject, subject.code, nota)

	def _apply_mp(self, student, subject, nota, ctx, stats):
		score, scored = self._coerce_score(nota)
		if not scored:
			# Textual MP (PQ/NP/...) has no official final to store; the state emerges on its own.
			return
		line = ctx["subject_line"].get((student.id, subject.id))
		if not line:
			self._log_error(stats, student.student_id, subject.code, "MP",
				_("No subject grade line for student (not enrolled or session not filled)."))
			return
		external_ponderation = line.grade_session_id.planning_id.external_ponderation
		if not external_ponderation:
			# No work placement: final == internal, so the official MP is stored exactly as an override.
			line.write({"is_overridden": True, "internal_score": score})
			stats["mp"] += 1
			self._log(stats, "MP", student, subject, subject.code, nota)
		else:
			# With work placement, EMS recomputes the final from RA + EM; only warn if it diverges.
			line.invalidate_recordset(["final_score"])
			if line.final_score != score:
				stats["warnings"].append(
					_("Module final mismatch for %s / %s: file says %s, EMS computes %s.")
					% (student.display_name, subject.code, score, line.final_score))
			self._log(stats, "MP", student, subject, subject.code, nota)

	@api.model
	def _coerce_score(self, value):
		# Returns (score, is_scored). Non-numeric notes (PDT/NA/NP/PQ) are "not scored".
		if value is None or value == "":
			return 0, False
		try:
			score = int(round(float(value)))
		except (TypeError, ValueError):
			return 0, False
		return max(0, min(10, score)), True

	# --- Logging and result ---

	def _log(self, stats, tipus, student, subject, code, nota):
		stats["log"].append({
			"tipus": tipus, "accio": "OK", "idalu": student.student_id,
			"alumne": student.display_name, "modul": subject.code, "codi": code,
			"nota": "" if nota is None else str(nota),
		})

	def _log_error(self, stats, idalu, modul, tipus, message):
		stats["errors"].append(message)
		stats["log"].append({
			"tipus": tipus, "accio": "ERROR", "idalu": idalu or "",
			"alumne": "", "modul": modul or "", "codi": "", "nota": message,
		})

	def _build_log_csv(self, log_entries):
		output = io.StringIO()
		writer = csv.writer(output)
		writer.writerow(["tipus", "accio", "idAlumne", "alumne", "modul", "codi", "nota / missatge"])
		for entry in log_entries:
			writer.writerow([
				entry["tipus"], entry["accio"], entry["idalu"], entry["alumne"],
				entry["modul"], entry["codi"], entry["nota"],
			])
		return base64.b64encode(output.getvalue().encode("utf-8-sig")).decode()

	def _build_result_html(self, stats):
		warnings_html = ""
		if stats["warnings"]:
			items = "".join("<li>%s</li>" % w for w in stats["warnings"])
			warnings_html = "<p><strong>%s (%d):</strong></p><ul>%s</ul>" % (_("Warnings"), len(stats["warnings"]), items)
		errors_html = ""
		if stats["errors"]:
			items = "".join("<li>%s</li>" % e for e in stats["errors"])
			errors_html = "<p><strong>%s (%d):</strong></p><ul>%s</ul>" % (_("Errors"), len(stats["errors"]), items)
		enrollments_html = ""
		if stats["enrollments"]:
			enrollments_html = "<p>📝 <strong>%s:</strong> %d</p>" % (
				_("Missing enrollments created"), stats["enrollments"])
		return (
			"<p>✅ <strong>%s:</strong> %d</p>"
			"<p>🏢 <strong>%s:</strong> %d</p>"
			"<p>📊 <strong>%s:</strong> %d</p>"
			"<p>🔓 <strong>%s:</strong> %d</p>"
			"%s%s%s"
		) % (
			_("Learning outcome grades applied"), stats["ra"],
			_("Work placement grades applied"), stats["em"],
			_("Module final grades overridden"), stats["mp"],
			_("Locked outcomes overwritten"), stats["locked"],
			enrollments_html, warnings_html, errors_html,
		)
